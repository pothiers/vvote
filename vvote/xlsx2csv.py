#! /usr/bin/env python
"""Convert Excel to CSV using openpyxl.
Other options:
- Gnumeric comes with "ssconvert"
- libreoffice --headless --convert-to csv $filename(s) --outdir $outdir


"""
# Docstrings intended for document generation via pydoc

import sys
import argparse
import logging
from openpyxl import load_workbook
from openpyxl import Workbook
import csv

def xlsx2csv(xlsx_filename, csv_filename,
             verbose=True, transpose=False, nrows=10000):
    if verbose:
        print('# Output status every {} rows'.format(nrows))
    wb = load_workbook(filename=xlsx_filename)
    ws0 = wb.active
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = 'Transposed'
    if (ws0.max_row == 1) or (ws0.max_column == 1):
        ws0.max_row = ws0.max_column = None
        # unzip -p /data/mock-election/Final_Count_LVR.xlsx | grep dimension
        ws0.calculate_dimension(force=True)
    print('# maxCol={}, maxRow={}'.format(ws0.max_column, ws0.max_row))

    if transpose:
        for row in range(1,ws0.max_row+1):
            for col in range(1,ws0.max_column+1):
                ws2.cell(row=col, column=row,
                         value=ws0.cell(row=row, column=col).value)
        ws = ws2
    else:
        ws = ws0        
                
    with open(csv_filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile, dialect='unix')
        ridx = 0
        for row in ws.rows:
            if verbose:
                if (ridx % nrows) == 0:
                    print('# processed {} ballots'.format(ridx))
            ridx += 1
            writer.writerow([cell.value for cell in row])


##############################################################################

def main():
    "Parse command line arguments and do the work."
    #!print('EXECUTING: %s\n\n' % (' '.join(sys.argv)))
    parser = argparse.ArgumentParser(
        description='Convert Excel (.xslx) to CSV (Comma Prated Values)',
        epilog='EXAMPLE: %(prog)s in.xslx out.csv'
        )
    parser.add_argument('--version', action='version', version='1.0.1')
    parser.add_argument('xlsxfile', # type=argparse.FileType('rb'),
                        help='Excel Input filename')
    parser.add_argument('csvfile', # type=argparse.FileType('w'),
                        help='CSV output filename')
    parser.add_argument('-t', '--transpose',
                        action='store_true',
                        help='Tranpose rows/columns on write to csvfile')
    parser.add_argument('--loglevel',
                        help='Kind of diagnostic output',
                        choices=['CRTICAL', 'ERROR', 'WARNING',
                                 'INFO', 'DEBUG'],
                        default='WARNING')
    args = parser.parse_args()
    #!args.csvfile.close()
    #!args.csvfile = args.csvfile.name


    log_level = getattr(logging, args.loglevel.upper(), None)
    if not isinstance(log_level, int):
        parser.error('Invalid log level: %s' % args.loglevel)
    logging.basicConfig(level=log_level,
                        format='%(levelname)s %(message)s',
                        datefmt='%m-%d %H:%M')
    logging.debug('Debug output is enabled in %s !!!', sys.argv[0])

    xlsx2csv(args.xlsxfile, args.csvfile,
             transpose=args.transpose,
             verbose=True,
             nrows=10000)

if __name__ == '__main__':
    main()
