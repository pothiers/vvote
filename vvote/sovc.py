#! /usr/bin/env python3
"""Extract from Statement Of Votes Cast (SOVC)
"""

import sys
import argparse
import logging

from collections import defaultdict
from pprint import pprint

from openpyxl import load_workbook
from . import exceptions as ex
from .db_ballot import BallotDb


class Sovc():
    """Maintain Statement Of Votes Cast (SOVC). Its an excel (.xslx) file
that represents official results.

   Row 1:: Race titles (duplicated over columns representing choices)
   Row 2:: party (we don't care)
   Row 3:: Choices
   Row 4 to N-1:: Precinct totals
   Row N:: Grand totals (County totals)
   Row N+1:: "_x001A_"  ??? End of data?

   Col 1:: County Number ('_x001A_' in last row)
   Col 2:: Precinct Code (number)
   Col 3:: Precinct Name (number) or "COUNTY TOTALS"
   Col 4:: "REGISTERED VOTERS - TOTAL" (Row 1)
   Col 5:: Ballots Cast-Total
   Col 6:: Ballots Cast-Blank
   Col 7 to M:: vote counts
"""
    MARKER='_x001A_'
    MARKER2='REGISTERED VOTERS - TOTAL'

    
    def __init__(self, sovc_file,
                 nrows = 10000, # progress every N rows iff verbose==True
                 verbose=False):
        wb = load_workbook(filename=sovc_file, read_only=True)
        ws = wb.active
        if (ws.max_row == 1) or (ws.max_column == 1):
            ws.max_row = ws.max_column = None
            ws.calculate_dimension(force=True)
        totalsrow = ws.max_row - 1        
        if verbose:
            print('DBG: file={}, ws.max_row = {}, ws.max_column = {}'
                  .format(sovc_file, ws.max_row, ws.max_column))

        # Validate format/content
        if ws.cell(row=1, column=4).value.strip() != self.MARKER2:
            msg = ('Row={}, Col={} is "{}" but expected "{}"'
                   .format(1, 4,
                           ws.cell(row=1, column=4).value,
                           self.MARKER2 ))
            raise ex.BadSovc('Invalid SOVC ({}); {}'.format(sovc_file, msg))

        if ws.cell(row=totalsrow+1, column=1).value.strip() != self.MARKER:
            msg = ('Row={}, Col={} is "{}" but expected "{}"'
                   .format(totalsrow, 1,
                           ws.cell(row=totalsrow+1, column=1).value,
                           self.MARKER ))
            raise  ex.BadSovc('Invalid SOVC ({}); {}'.format(sovc_file, msg))
        if ws.cell(row=totalsrow, column=3).value != 'COUNTY TOTALS':
            msg = ('Row={}, Col={} is "{}" but expected "COUNTY TOTALS"'
                   .format(totalsrow, 3,
                           ws.cell(row=totalsrow, column=3).value))
            raise ex.BadSovc('Invalid SOVC ({}); {}'.format(sovc_file, msg))

        self.filename = sovc_file
        self.totalsrow = totalsrow
        self.ws = ws
        self.max_row = ws.max_row
        self.max_column = ws.max_column

    def get_totals(self):
        "RETURN: dict[(race,choice)] => count"
        races = [self.ws.cell(row=1, column=c).value.strip()
                 for c in range(4, self.max_column+1)]
        choices = [self.ws.cell(row=3, column=c).value.strip()
                   for c in range(4, self.max_column+1)]
        totals = [self.ws.cell(row=self.totalsrow, column=c).value
                  for c in range(4, self.max_column+1)]
        racechoice = zip(races, choices)
        totdict = dict(zip(racechoice, totals))
        return totdict

    def get_precinct_totals(self):
        "RETURN: dict[(race,choice)] => (count,precinct,regvot,baltot,balblank)"
        races = [self.ws.cell(row=1, column=c).value.strip()
                 for c in range(4, self.max_column+1)]
        choices = [self.ws.cell(row=3, column=c).value.strip()
                   for c in range(4, self.max_column+1)]
        totdict = dict()
        #!for r in range(4, self.totalsrow):
        #!for r,row in enumerate(self.ws.rows, start=1):
        #!for r,row in enumerate(self.ws.rows[4:], start=4):
        for row in list(self.ws.rows)[4:]:
            (county,pcode,precinct,numreg,btotal,bblank,*tally) = row
            print('Save data for precinct_code={}'.format(pcode.value))
            for cell in tally:
                if cell.value == None: continue
                race = self.ws.cell(row=1, column=cell.column).value
                choice = self.ws.cell(row=3, column=cell.column).value
                totdict[(race.strip(), choice.strip())] = (
                    cell.value,
                    precinct.value, numreg.value, btotal.value, bblank.value)
        print('DBG-5')
        return totdict

    def get_races(self):
        return set([self.ws.cell(row=1, column=c).value.strip()
                   for c in range(7, self.max_column+1)])

    def get_choices(self):
        other = ['BALLOTS CAST', 'OVER VOTES', 'UNDER VOTES',
                 'VOTERS', 'WRITE-IN']
        choices = set([self.ws.cell(row=3, column=c).value.strip()
                       for c in range(4, self.max_column+1)])
        #!return choices - set(other)
        return choices

    def get_race_choices(self, race):
        choices = set()
        for c in range(4, self.max_column+1):
            racetitle = self.ws.cell(row=1, column=c).value.strip()
            if race == racetitle:
                choices.add(self.ws.cell(row=3, column=c).value.strip())
        return choices


    def save(self, dbfile='SOVC.db'):
        """Save contents in sqlite database"""
        sovcdb = BallotDb(dbfile, self.filename)
        
        print('DBG: save RACE and CHOICE tables')
        cid = 0
        choiceLut = dict() # lut[title] = id
        raceLut = dict() # lut[title] = id
        race_list = list()
        choice_list = list()
        for c in range(4, self.max_column+1):
            racetitle = self.ws.cell(row=1, column=c).value.strip()
            rid = c
            race_list.append((rid, racetitle, None))
            raceLut[racetitle] = rid
            for c2 in range(c, self.max_column+1):
                if racetitle == self.ws.cell(row=1, column=c2).value.strip():
                    choicetitle = self.ws.cell(row=3, column=c2).value.strip()
                    choice_list.append((cid, choicetitle, rid, None))
                    choiceLut[choicetitle] = cid
                    cid += 1
                else:
                    break
        sovcdb.insert_race_list(race_list)
        sovcdb.insert_choice_list(choice_list)

        print('DBG: save PRECINCT table')
        precinct_list = list()
        vote_list = list()
        # dict[(race,choice] => (count,precinct,regvot,baltot,balblank)"
        pt = self.get_precinct_totals()
        for ((racetitle,choicetitle),
             (count,precinct,regvot,baltot,balblank)) in pt.items():
            choice_id = choiceLut.get(choicetitle, None)
            precinct_list.append((raceLut[racetitle],
                                  choice_id,
                                  None, # county_number
                                  precinct,
                                  precinct,
                                  regvot, # registered_voters integer,
                                  baltot, # ballots_cast_total integer,
                                  balblank # ballots_cast_blank integer
                                  ))
            vote_list.append((choice_id, precinct, count))
        sovcdb.insert_precinct_list(precinct_list)        
        sovcdb.insert_vote_list(vote_list)        
        sovcdb.close()
        print('Saved SOVC to sqlite DB: {}'.format(dbfile))

##############################################################################

def main():
    "Parse command line arguments and do the work."
    #print('EXECUTING: %s\n\n' % (' '.join(sys.argv)))
    parser = argparse.ArgumentParser(
        description='Extract from Statement Of Votes Cast (SOVC) excel (.xslx)',
        epilog='EXAMPLE: %(prog)s sovc.xslx'
        )
    parser.add_argument('--version', action='version', version='1.0.1')
    parser.add_argument('infile', type=argparse.FileType('r'),
                        help='Input file')
    #!parser.add_argument('outfile', type=argparse.FileType('w'),
    #!                    help='Vote count output')

    parser.add_argument('--csv',
                        action='store_true',
                        help='Write Excel to CSV file')

    parser.add_argument('--loglevel',
                        help='Kind of diagnostic output',
                        choices=['CRTICAL', 'ERROR', 'WARNING',
                                 'INFO', 'DEBUG'],
                        default='WARNING')
    args = parser.parse_args()
    args.infile.close()
    args.infile = args.infile.name
    #args.outfile.close()
    #args.outfile = args.outfile.name

    log_level = getattr(logging, args.loglevel.upper(), None)
    if not isinstance(log_level, int):
        parser.error('Invalid log level: %s' % args.loglevel)
    logging.basicConfig(level=log_level,
                        format='%(levelname)s %(message)s',
                        datefmt='%m-%d %H:%M')
    logging.debug('Debug output is enabled in %s !!!', sys.argv[0])

    #!print('Reading counts from file: "{}"'.format(args.infile))
    sovc = Sovc(args.infile)
    print('Saving to DB')
    sovc.save()
    #!print('Getting totals')
    #!totdict = sovc.get_totals()
    #!print('Totals = ',)
    #!pprint(totdict)

if __name__ == '__main__':
    main()
