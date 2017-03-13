#! /usr/bin/env python3
"""Extract from Cast Vote Record (CVR)
"""

from openpyxl import load_workbook

def valid_CVR(cvr_file, verbose=False):
    # Row 1:: Race titles (None over columns representing choices beyond first)
    # Row 2 to N:: Choices
    #
    # Col 4 to M::

    MARKER='Cast Vote Record'
    wb = load_workbook(filename=cvr_file, read_only=True)
    ws = wb.active
    if (ws.max_row == 1) or (ws.max_column == 1):
        ws.max_row = ws.max_column = None
        ws.calculate_dimension(force=True)
    totalsrow = ws.max_row - 1        
    if verbose:
        print('DBG: file={}, ws.max_row = {}, ws.max_column = {}'
              .format(cvr_file, ws.max_row, ws.max_column))

    if ws.cell(row=1, column=1).value.strip() != MARKER:
        msg = ('Row={}, Col={} is "{}" but expected "{}"'
               .format(totalsrow, 1, ws.cell(row=totalsrow+1, column=1).value,
                       MARKER ))
        raise 'Invalid CVR ({}); {}'.format(cvr_file, msg)
    return ws

