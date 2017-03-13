#! /usr/bin/env python3
"""Count ballots.
"""

import sys
import argparse
import logging
import sqlite3

from openpyxl import load_workbook
from openpyxl import Workbook
from .sovc import Sovc
from .lvr import Lvr
from . import excel_utils as eu


#!ballot2sovc = {
#!    #Ballot, SOVC
#!    'Write-in': 'WRITE-IN',
#!    'overvote': 'OVER VOTES',
#!    'undervote': 'UNDER VOTES',
#!    'YES/SÍ': 'YES/SI' ,
#!    'PRESIDENTIAL ELECTOR': 'PRESIDENTIAL ELECTORS',
#!    'U.S. SENATOR': 'UNITED STATES SENATOR',
#!    }
#!
#!def revlut(lut):
#!    "RETURN: new dict[a] => b from dict[b] => a"
#!    newlut = dict()
#!    for k,v in lut.items():
#!        newlut[v] = k
#!    return newlut
#!
#!sovc2ballot = revlut(ballot2sovc)

            


def emit_results(votes, choices, n_votes, num_ballots,
                 orderedraces=None,
                 na_tag='<OOD>', # Out of District Ballots
                 outputfile = None):
    # votes[race][choice] = count
    # choices[race] = set([choice1, choice2,...])
    # n_votes[race] => num-to-vote-for

    if outputfile == None:
        file=sys.stdout
    else:
        file = open(outputfile, mode='w')

    for race in orderedraces:
        print(file=file)
        print("{} (vote for {})".format(race, n_votes[race]), file=file)
        #for choice in sorted(choices[race]):
        special = set([k for k in votes[race].keys()
                       if k[:10] == 'undervote'])
        for choice in sorted(set(votes[race].keys()) - special):
            #!if choice == na_tag:
            #!    n = n_votes[race]
            #!    print('\t{}\t{}'.format(choice, int(votes[race][choice]/n)),
            #!          file=file)
            if choice != na_tag:
                print('\t{:10d}\t{}'.format(votes[race][choice], choice),
                      file=file)
        print('\t{:10d}\t{}'.format(num_ballots - votes[race][na_tag],
                                'IN-DISTRICT'),
              file=file)
        print('\t{:10d}\t{}'.format(votes[race][na_tag], 'OUT-OF-DISTRICT'),
              file=file)
    file.close()


schema = '''
CREATE TABLE race (
   race_id integer primary key,
   title text,
   num_to_vote_for integer,
);
CREATE TABLE choices (
  race_id integer,
  choice_id text,
);
CREATE TABLE cvr (
  cvr_id integer,
  race_id integer,
  choice_id text,
);

'''    
#!def save_sovc_db(dbfile, votes, choices, n_votes, sovcfilename,
#!                 orderedraces=None,
#!                 na_tag='<OOD>'):
#!    """Save SOVC sqlite DB"""
#!    # votes[race][choice] = count
#!    # choices[race] = set([choice1, choice2,...])
#!    # n_votes[race] => num-to-vote-for
#!
#!    conn = sqlite3.connect(dbfile)
#!
#!    for race in orderedraces:
#!        nvotes = n_votes[race]
#!
#!        # sum values from N*undervote-N (N=[1..nvotes])
#!        undervotes = 0
#!        for n in range(1,nvotes+1):
#!            choice = 'undervote-{}'.format(n)
#!            undervotes += (n * votes[race][choice])
#!            ignore_choices.add(choice) 
#!        votes[race]['undervotes'] = undervotes
#!        choices[race].add('undervotes')
#!        
#!        for choice in set(choices[race]-ignore_choices):
#!            count = votes[race][choice]
#!            if choice == 'overvote':
#!                count*= nvotes
#!
#!            ws.cell(column=col, row=1, value="{}".format(race))
#!            ws.cell(column=col, row=3, value="{}".format(choice))
#!            ws.cell(column=col, row=4, value="{}".format(count))
#!            col += 1
#!    wb.save(sovcfilename)
#!    sovc.transpose(sovcfilename, '{}.transpose.xlsx'.format(sovcfilename))


def write_sovc(votes, choices, n_votes, sovcfilename,
               orderedraces=None,
               orderedchoices=None,
               na_tag='<OOD>'):
    # votes[race][choice] = count
    # choices[race] = set([choice1, choice2,...])
    # n_votes[race] => num-to-vote-for

    # Row 1:: Race titles (merged over columns representing choices)
    # Row 2:: party (leave blank, we do not know)
    # Row 3:: Choices
    # Row 4:: Grand totals (County totals)

    wb = Workbook()
    ws = wb.active

    # Races
    ws['A1'] = 'COUNTY NUMBER'
    ws['B1'] = 'PRECINCT CODE'
    ws['C1'] = 'PRECINCT NAME'
    ws['D1'] = 'REGISTERED VOTERS - TOTAL'
    ws['E1'] = 'BALLOTS CAST - TOTAL'
    ws['F1'] = 'BALLOTS CAST - BLANK'
    # G1 ... :: Races (duplicated for each choice of same race)

    # Choices
    ws['D3'] = 'VOTERS'
    ws['E3'] = 'BALLOTS CAST'
    ws['F3'] = 'BALLOTS CAST'
    # G3 ... :: Candidates

    ws['B4'] = 'ZZZ'
    ws['C4'] = 'COUNTY TOTALS'
    ws['A5'] = '_x001A_'

    col = 7
    ignore_choices = set([na_tag])
    for race in orderedraces:
        nvotes = n_votes[race]

        # sum values from N*undervote-N (N=[1..nvotes])
        undervotes = 0
        for n in range(1,nvotes+1):
            choice = 'undervote-{}'.format(n)
            undervotes += (n * votes[race][choice])
            ignore_choices.add(choice) 
        votes[race]['undervotes'] = undervotes
        choices[race].add('undervotes')
        
        #for choice in set(choices[race]-ignore_choices):
        for choice in orderedchoices[race]:
            if choice in ignore_choices:
                continue
            count = votes[race][choice]
            if choice == 'overvote':
                count*= nvotes

            ws.cell(column=col, row=1, value="{}".format(race))
            ws.cell(column=col, row=3, value="{}".format(choice))
            ws.cell(column=col, row=4, value="{}".format(count))
            col += 1
    wb.save(sovcfilename)
    eu.transpose(sovcfilename, '{}.transpose.xlsx'.format(sovcfilename))


def clean_choice(choice):
    "Systemic fix to strings given as choices in BALLOTS"
    if choice[:4] == 'DEM ':
        return choice[4:]
    if choice[:4] == 'REP ':
        return choice[4:]
    return choice.strip()
    



##############################################################################

def main():
    "Parse command line arguments and do the work."
    #print('EXECUTING: %s\n\n' % (' '.join(sys.argv)))
    parser = argparse.ArgumentParser(
        description='Count ballots.',
        epilog='EXAMPLE: %(prog)s --verbose --sovc /data/mock-election/G2016_EXPORT1.xlsx /data/mock-election/day-1-cvr.xlsx results1.txt'
        #  (49418 ballots)
        )
    parser.add_argument('--version', action='version', version='1.0.1')
    parser.add_argument('LVRfile', type=argparse.FileType('r'),
                        help='Input LVR excel (.xslx) file')
    parser.add_argument('outfile', type=argparse.FileType('w'),
                        help='Vote count output (SOVC equiv)')

    parser.add_argument('-s', '--sovc', type=argparse.FileType('r'),
                        help=('Name of SOVC (xslx) file to compare to results.'
                              '(No comparison done if not given.)'),
                        )
    parser.add_argument('-f', '--format',
                        help='Format of output file.',
                        choices=['text', 'SOVC'],
                        default='text')
    parser.add_argument('-v', '--verbose',
                        action='store_true',
                        help='Output progress')

    parser.add_argument('--loglevel',
                        help='Kind of diagnostic output',
                        choices=['CRTICAL', 'ERROR', 'WARNING',
                                 'INFO', 'DEBUG'],
                        default='WARNING')
    args = parser.parse_args()
    args.LVRfile.close()
    args.LVRfile = args.LVRfile.name
    args.outfile.close()
    args.outfile = args.outfile.name
    if args.sovc:
        args.sovc.close()
        args.sovc = args.sovc.name

    log_level = getattr(logging, args.loglevel.upper(), None)
    if not isinstance(log_level, int):
        parser.error('Invalid log level: %s' % args.loglevel)
    logging.basicConfig(level=log_level,
                        format='%(levelname)s %(message)s',
                        datefmt='%m-%d %H:%M')
    logging.debug('Debug output is enabled in %s !!!', sys.argv[0])

    print('# Counting votes from file: {}'.format(args.LVRfile))
    lvr = Lvr(args.LVRfile)
    (votes, choices, n_votes, nballots, races, ochoices
    ) = lvr.count_votes(verbose=args.verbose)
    # Vote counts now in: votes
    if args.format == 'text':
        emit_results(votes, choices, n_votes, nballots,
                     orderedraces=races,
                     outputfile=args.outfile )
    elif args.format == 'SOVC':
        write_sovc(votes, choices, n_votes, args.outfile,
                   orderedraces=races, orderedchoices=ochoices)
        
    if args.sovc != None:
        sovc = Sovc(args.sovc)
        sovc.compare(votes, choices, n_votes)

if __name__ == '__main__':
    main()
