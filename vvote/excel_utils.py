
from openpyxl import load_workbook
from openpyxl import Workbook


def transpose(in_xslx, out_xslx):
    "Swap rows/columns"
    wb = load_workbook(filename=in_xslx)
    ws = wb.active
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = 'Transposed'

    for row in range(1, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            ws2.cell(row=col, column=row,
                     value=ws.cell(row=row, column=col).value)
    wb2.save(filename=out_xslx)
    
