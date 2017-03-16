#! /usr/bin/env python3
"""Extract from List of Vote Records (LVR)
"""

import copy
import sys
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl import Workbook

from . import excel_utils as eu


class Lvr():
    """Maintain List of Vote Records (LVR). Its an excel (.xslx) file
that represents contents from a set of ballots.

   Row 1:: Race titles 
   Row 2 to N:: Choices

   Col 1:: CVR id
   Col 2:: Precinct
   Col 3:: Ballot Style
   Col 4 to M::
"""
    # votes[(race,precinct)][choice] = cnt; Precinct is number or "ALL"
    votes = defaultdict(lambda : defaultdict(int)) 
    n_votes = defaultdict(int)  # n_votes[race] => num-to-vote-for
    choices = defaultdict(set) # choices[race] = set([choice1, choice2,...])
    orderedchoices = dict() # d[race] => [choice1, ...]
    orderedraces = list()
    precincts = set('ALL')
    nballots = 0

    MARKER='Cast Vote Record'
    nontitles = set(['Cast Vote Record',
                     'Serial Number',
                     'Precinct',
                     'Ballot Style'])

    def __init__(self, lvr_file, # Excel (.xslx) filename,
                 verbose=False):
        wb = load_workbook(filename=lvr_file, read_only=True)
        ws = wb.active
        if (ws.max_row == 1) or (ws.max_column == 1):
            ws.max_row = ws.max_column = None
            ws.calculate_dimension(force=True)
        totalsrow = ws.max_row - 1        
        if verbose:
            print('DBG: file={}, ws.max_row = {}, ws.max_column = {}'
                  .format(lvr_file, ws.max_row, ws.max_column))

        if ws.cell(row=1, column=1).value.strip() != self.MARKER:
            msg = ('Row={}, Col={} is "{}" but expected "{}"'
                   .format(totalsrow, 1,
                           ws.cell(row=1, column=1).value,
                           self.MARKER ))
            raise 'Invalid LVR ({}); {}'.format(lvr_file, msg)

        self.filename = lvr_file
        self.ws = ws
        self.max_row = ws.max_row
        self.max_column = ws.max_column



    def count_votes(self,
                    verbose=False,
                    nrows = 10000, # progress every N rows iff verbose==True
                    na_tag='<OOD>', # Out of District Ballots
                    writeintag='Write-in',
                    overvotetag='overvote',
                    undervotetag='undervote'):
        """Talley ballots. Store results in class variables."""

        votes = self.votes
        choices = self.choices
        n_votes = self.n_votes
        num_ballots = self.nballots
        orderedraces = self.orderedraces
        orderedchoices = self.orderedchoices

        coltitle = dict() # coltitle[column] => racetitle
        raceballot = list() # single ballot for single race, list(choice1, ...])

        prevtitle = None
        ridx = 0
        ws = self.ws
        precinct_col = None
        for row in ws.rows:
            ridx += 1
            cidx = 0
            if verbose:
                if (ridx % nrows) == 0:
                    print('# processed {} ballots'.format(ridx))
            for cell in row:
                cidx += 1
                #print('# DBG-0: ridx={} cidx{}'.format(ridx,cidx))

                if ridx == 1 and cell.value == 'Precinct':
                    precinct_col = cidx
                    
                # Ignore the (leading) columns that are not Race Titles
                if ridx == 1 and cell.value in self.nontitles:
                    continue

                if ridx == 1: # header
                    if cell.value != None:
                        race = cell.value.strip()
                        orderedraces.append(race)
                        n_votes[race] = 1
                        #!choices[race].add(overvotetag)
                        #!choices[race].add(writeintag)
                        #!choices[race].add(na_tag)
                    else: # vote-for-N race
                        n_votes[race] += 1
                    coltitle[cidx] = race
                else: # ballots
                    if cidx == precinct_col:
                        precinct = cell.value
                        self.precincts.add(precinct)
                    if cidx not in coltitle:
                        # Skip columns we don't care about
                        continue

                    if (cell.value == '' or cell.value == None):
                        choice = na_tag
                    else:
                        #choice = clean_choice(cell.value)
                        choice = cell.value
                    race = coltitle[cidx]
                    next_race = coltitle[cidx+1] if cidx < ws.max_column else None
                    #print('# DBG-0: race="{}"'.format(race))
                    raceballot.append(choice)
                    if race != next_race: # finished one race, one ballot
                        #print('DBG-1: raceballot={}'.format(raceballot))
                        if undervotetag in raceballot:
                            undervote_m = ('undervote-{}'
                                           .format(raceballot.count(undervotetag)))
                            votes[(race,precinct)][undervote_m] += 1
                            votes[(race,'ALL')][undervote_m] += 1
                            #!choices[race].add(undervote_m)
                            raceballot.append(undervote_m)
                        if overvotetag in raceballot:
                            assert raceballot.count(overvotetag) == n_votes[race]
                            votes[(race,precinct)][overvotetag] += 1
                            votes[(race,'ALL')][overvotetag] += 1
                        if na_tag in raceballot:
                            assert raceballot.count(na_tag) == n_votes[race]
                            votes[(race,precinct)][na_tag] += 1
                            votes[(race,'ALL')][na_tag] += 1
                        for c in raceballot:
                            if c == writeintag:
                                #!choices[race].add(c)
                                votes[(race,precinct)][c] += 1          
                                votes[(race,'ALL')][c] += 1          

                        raceballot = [c for c in raceballot
                                      if ((c != undervotetag)
                                          and (c != overvotetag)
                                          and (c != na_tag)
                                          and (c != writeintag)
                                      )]
                        # no dupes
                        assert len(raceballot) == len(set(raceballot)), 'Duplicates in: {}'.format(raceballot)

                        for choice in set(raceballot):
                            choices[race].add(choice)
                            votes[(race,precinct)][choice] += 1
                            votes[(race,'ALL')][choice] += 1
                        orderedchoices[race] = copy.copy(raceballot)
                        raceballot = list() # single ballot for single race

        self.votes = votes
        self.choices = choices
        self.n_votes = n_votes
        self.nballots = ridx-1
        self.orderedraces = orderedraces
        self.orderedchoices = orderedchoices
        print('Processed {} ballots'.format(self.nballots))


    def emit_results(self, totals_only=False, outputfile = None):
        na_tag='<OOD>' # Out of District Ballots
        votes = self.votes
        choices = self.choices
        n_votes = self.n_votes
        num_ballots = self.nballots
        orderedraces = self.orderedraces

        if outputfile == None:
            file=sys.stdout
        else:
            file = open(outputfile, mode='w')
        plist = sorted(list((p for p in self.precincts if p != 'ALL')))
        preclist = ['ALL'] if totals_only else (plist + ['ALL'])
        for prec in preclist:
            for race in orderedraces:
                rp = (race,prec) 
                print(file=file)
                print("{}, precint:{} (vote for {})"
                      .format(race, prec, n_votes[race]), file=file)
                #for choice in sorted(choices[race]):
                special = set([k for k in votes[rp].keys()
                               if k[:10] == 'undervote'])
                for choice in sorted(set(votes[rp].keys()) - special):
                    if choice != na_tag:
                        print('\t{:10d}\t{}'
                              .format(votes[rp][choice], choice),
                              file=file)
                print('\t{:10d}\t{}'
                      .format(num_ballots - votes[rp][na_tag], 'IN-DISTRICT'),
                      file=file)
                print('\t{:10d}\t{}'
                      .format(votes[rp][na_tag], 'OUT-OF-DISTRICT'),
                      file=file)
        file.close()


    def write_sovc(self,  sovcfilename,
                   orderedchoices=None,
                   na_tag='<OOD>'):
        votes = self.votes
        choices = self.choices
        n_votes = self.n_votes
        num_ballots = self.nballots
        orderedraces = self.orderedraces
        orderedchoices = self.orderedchoices
        
        # votes[race][choice] = count
        # choices[race] = set([choice1, choice2,...])
        # n_votes[race] => num-to-vote-for

        # Row 1:: Race titles (merged over columns representing choices)
        # Row 2:: party (leave blank, we do not know)
        # Row 3:: Choices
        # Row 4:: Grand totals (County totals)

        wb = Workbook()
        ws = wb.active

        # Races
        ws['A1'] = 'COUNTY NUMBER'
        ws['B1'] = 'PRECINCT CODE'
        ws['C1'] = 'PRECINCT NAME'
        ws['D1'] = 'REGISTERED VOTERS - TOTAL'
        ws['E1'] = 'BALLOTS CAST - TOTAL'
        ws['F1'] = 'BALLOTS CAST - BLANK'
        # G1 ... :: Races (duplicated for each choice of same race)

        # Choices
        ws['D3'] = 'VOTERS'
        ws['E3'] = 'BALLOTS CAST'
        ws['F3'] = 'BALLOTS CAST'
        # G3 ... :: Candidates

        ws['B4'] = 'ZZZ'
        ws['C4'] = 'COUNTY TOTALS'

        col = 7
        ignore_choices = set([na_tag])
        for race in orderedraces:
            nvotes = n_votes[race]

            # sum values from N*undervote-N (N=[1..nvotes])
            undervotes = 0
            for n in range(1,nvotes+1):
                choice = 'undervote-{}'.format(n)
                undervotes += (n * votes[race][choice])
                ignore_choices.add(choice) 
            votes[race]['undervotes'] = undervotes
            choices[race].add('undervotes')

            #for choice in set(choices[race]-ignore_choices):
            for choice in orderedchoices[race]:
                if choice in ignore_choices:
                    continue
                ws.cell(column=col, row=1).value = race
                ws.cell(column=col, row=3).value = choice
                #!count = votes[race][choice]
                #!if choice == 'overvote':
                #!    count *= nvotes
                #!ws.cell(column=col, row=4).value = count
                r=4
                prec_list = sorted(self.precincts - {'ALL'})
                for prec in prec_list:
                    rp = (race,prec) 
                    count = votes[rp][choice]
                    if choice == 'overvote':
                        count *= nvotes
                    ws.cell(column=col, row=r).value = count
                    ws.cell(column=2, row=r).value = prec
                    ws.cell(column=3, row=r).value = prec
                    r += 1
                prec = 'ALL'
                rp = (race,prec) 
                count = votes[rp][choice]
                if choice == 'overvote':
                    count *= nvotes
                ws.cell(column=col, row=r).value = count
                ws.cell(column=2, row=r).value = prec
                ws.cell(column=3, row=r).value = prec
                ws.cell(column=1, row=r+1).value = '_x001A_'
                col += 1
        wb.save(sovcfilename)
        eu.transpose(sovcfilename, '{}.transpose.xlsx'.format(sovcfilename))

        
    def get_titles(self,
                   nrows = 10000, # progress every N rows iff verbose==True
                   verbose=False):
        "RETURN: dict[(race,choice)] => count"

        ws = self.ws
        other = ['undervote', 'Write-in']
        races = set()
        choices = set()
        ignorecolumns = set()
        ridx = 0
        for row in ws.rows:
            ridx += 1
            cidx = 0
            if verbose:
                if (ridx % nrows) == 0:
                    print('# processed {} ballots'.format(ridx))
            for cell in row:
                cidx += 1
                if cidx in ignorecolumns:
                    continue
                # Ignore the (leading) columns that are not Race Titles
                if ridx == 1 and cell.value in self.nontitles:
                    ignorecolumns.add(cidx)
                    continue
                if ridx == 1: # header
                    if cell.value != None:
                        races.add(cell.value.strip())
                else: #ballots
                    if (cell.value == '' or cell.value == None or cell.value.strip() == ''):
                        continue
                    choices.add(cell.value.strip())
        return races, (choices - set(other))
